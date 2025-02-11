# Bibliotecas utilizadas:
#  Pathlib para acessar as pastas 
from pathlib import Path
# Openpyxl para extrair dados do Excel e exportá-los em outro arquivo Excel
from openpyxl import load_workbook, Workbook

# Contém o caminho da pasta onde estão armazenados os ensaios
import endereco

# Define o caminho a ser utilizado
endereco = Path(endereco.pasta_ensaios)

# Formatos de arquivo que serão verificados
formatos = ['.xlsx', '.xlsm']

# Definições iniciais do arquivo Resumo em Excel
wb = Workbook()
resumo = wb.active
resumo.title = 'Resumo'
linha = 2
coluna = 1
wb.save('Resumo.xlsx')

# Função para percorrer as pastas dos ensaios
def percorrer_pastas(caminho):

    for pasta in caminho.iterdir():
        # Se for um arquivo, foram determinadas 3 condições: 
        # - Formato válido (.xlsx ou .xlsm)
        # - Acessa apenas arquivos do sistema de gestão (como padrão da empresa, todos os arquivos deste tipo têm FORM no nome)
        # - Não possui o '$' no nome para evitar a abertura de arquivos temporários no Excel, podendo ocasionar erros
        if pasta.suffix in formatos and 'FORM' in pasta.name and '$' not in pasta.name:
            try: 
                # Chama a função extrair_dados e seu retorno é utilizado como argumento para a função preencher_resumo
                preencher_resumo(extrair_dados(pasta)[0], extrair_dados(pasta)[1])
            except:
                continue
        # Se for uma pasta, chama a função novamente para encontrar arquivos dentro dela (função recursiva)    
        if pasta.is_dir():
            percorrer_pastas(pasta)

# Função para extrair os protocolos e equipamentos de cada arquivo Excel de ensaio
def extrair_dados(arquivo):
    equipamentos = []
    protocolos = []
    # Carrega o arquivo em Excel do ensaio
    wb = load_workbook(arquivo, data_only=True)
    
    # Itera sobre cada aba dentro do arquivo Excel
    for ws in wb.worksheets:
        # Itera sobre cada célula na aba linha a linha, dentro de um intervalo determinado
        for linha in ws.iter_rows(max_row=90, max_col=30):
            for celula in linha:
                formato = celula.number_format
                valor = celula.value
                # Condição: ignora células vazias
                if valor != None:
                    # Células com este formato possuem números de protocolo, que são identificações do material ensaiado
                    if formato == '000"/2024"' or formato == '000"/2025"':
                        # Evita a repetição dos números de protocolo ao extrair os dados
                        if valor not in protocolos:
                            protocolos.append(valor)
                    # Células com este formato possuem os equipamentos utilizados para o ensaio
                    if formato == '"LC"\ 000' and valor != 'LABORATÓRIO CENTRAL':
                        # Evita a repetição dos equipamentos ao extrair os dados
                        if valor not in equipamentos:
                            equipamentos.append(valor)
    
    # Se a planilha não possui preenchimento em nenhum campo de protocolo, ela é ignorada
    if protocolos != []:
        protocolo = protocolos[0]
        # Para ensaios realizados com mais de um material (pelo menos 2 protocolos diferentes), a saída será em um único valor no formato: 'PT 1 + PT 2 + ... PT N'
        if len(protocolos) > 1:
            for i in range(1,len(protocolos)):
                protocolo = str(protocolo) + ' + ' + str(protocolos[i])
        # Imprime os dados extraídos no console para verificar o andamento da execução do código 
        print(f'Protocolo: {protocolo}, Equipamentos: {equipamentos}')
        # A função retorna os dados extraídos: o número de protocolo e todos os equipamentos utilizados
        return protocolo, equipamentos

# Função para preencher os dados extraídos numa planilha de resumo em Excel
def preencher_resumo(protocolo, equipamentos):
    global coluna
    global linha
    # Cada coluna conterá os equipamentos de um protocolo
    coluna = coluna+1
    resumo.cell(row=linha, column=coluna).value = protocolo
    resumo.cell(row=linha, column=coluna).number_format = '000"/25"'
    # Para cada coluna, os equipamentos serão listados em linha
    for equipamento in equipamentos:
        linha=linha+1
        resumo.cell(row=linha, column=coluna).value = equipamento
        resumo.cell(row=linha, column=coluna).number_format = '"LC "000'
    #Reseta a linha para a próxima interação, que preencherá o novo protocolo na linha 2 e numa coluna à direita
    linha = 2
    
    wb.save('Resumo.xlsx')

percorrer_pastas(endereco)